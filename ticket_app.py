import logging
import pandas as pd
import os
import zipfile
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import smtplib
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput
from kivy.uix.checkbox import CheckBox
from kivy.uix.dropdown import DropDown
from kivy.uix.gridlayout import GridLayout
from kivy.graphics import Color, Rectangle
from kivy.core.window import Window
from kivy.uix.popup import Popup
from tkinter import Tk, messagebox
from tkinter.filedialog import askdirectory
import subprocess
import json
import base64
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
import requests
import shutil
import socket

# Set up logging
log_file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'app.log')
logging.basicConfig(filename=log_file_path, level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

# If modifying these SCOPES, delete the file token.json.
SCOPES = ['https://www.googleapis.com/auth/gmail.send']

def check_for_updates():
    try:
        # URL of the version file in the repository
        version_url = "https://raw.githubusercontent.com/jtester24/ANPCTechSupport/master/version.txt"
        
        # Get the current version from the repository
        response = requests.get(version_url)
        latest_version = response.text.strip()
        
        # Read the local version
        with open("version.txt", "r") as file:
            local_version = file.read().strip()
        
        # Compare versions
        if latest_version > local_version:
            messagebox.showinfo("Update Available", "A new version of ANPC Tech Support is available. The application will now update.")
            download_and_install_update(latest_version)
        else:
            messagebox.showinfo("No Update Available", "You are using the latest version of ANPC Tech Support.")
    except Exception as e:
        logging.error(f"Error checking for updates: {e}")

def download_and_install_update(latest_version):
    try:
        # URL of the installer in the repository
        installer_url = f"https://github.com/jtester24/ANPCTechSupport/releases/download/v{latest_version}/ANPCTechSupportInstaller.exe"
        
        # Download the installer
        installer_path = os.path.join(os.path.expanduser("~"), "Desktop", "ANPCTechSupportInstaller.exe")
        response = requests.get(installer_url)
        with open(installer_path, 'wb') as file:
            file.write(response.content)
        
        # Run the installer
        subprocess.run([installer_path], check=True)
        
        # Update the local version file
        with open("version.txt", "w") as file:
            file.write(latest_version)
        
        messagebox.showinfo("Update Complete", "ANPC Tech Support has been updated to the latest version.")
    except Exception as e:
        logging.error(f"Error downloading or installing the update: {e}")

def send_email(service, sender_email, recipients, subject, body, attachments=[]):
    message = MIMEMultipart()
    message['From'] = sender_email
    message['To'] = ", ".join(recipients)
    message['Subject'] = subject
    message.attach(MIMEText(body, 'html'))

    for attachment in attachments:
        with open(attachment, 'rb') as attachment_file:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment_file.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename={os.path.basename(attachment)}')
        message.attach(part)

    raw_message = base64.urlsafe_b64encode(message.as_bytes()).decode()
    message = {'raw': raw_message}

    try:
        message = (service.users().messages().send(userId="me", body=message).execute())
        logging.info(f"Message Id: {message['id']}")
        print(f"Message Id: {message['id']}")
    except Exception as e:
        logging.error(f"An error occurred: {e}")
        print(f"An error occurred: {e}")

class Tooltip(Popup):
    def __init__(self, title, text, **kwargs):
        super().__init__(**kwargs)
        self.size_hint = (None, None)
        self.size = (450, 225)  # Increased size by 50%
        self.auto_dismiss = True
        self.background = 'atlas://data/images/defaulttheme/button'
        self.content = BoxLayout(orientation='vertical', padding=10, spacing=5)
        self.content.add_widget(Label(text=title, font_size='20sp', bold=True, color=(1, 1, 1, 1)))
        label = Label(text=text, font_size='16sp', color=(1, 1, 1, 1), text_size=(420, None), size_hint_y=None)
        label.bind(texture_size=label.setter('size'))
        self.content.add_widget(label)

class ANPCSupportTicket(App):
    def build(self):
        # Check for updates when the application starts
        check_for_updates()
        
        # Create the main layout
        self.layout = BoxLayout(orientation='vertical', padding=20, spacing=20)
        
        # Set background color
        with self.layout.canvas.before:
            Color(0, 0, 0.5, 1)  # Dark blue color
            self.rect = Rectangle(size=(Window.width, Window.height), pos=self.layout.pos)
            self.layout.bind(size=self._update_rect, pos=self._update_rect)

        # Add title
        title = Label(text='ANPC Tech Support Ticket', font_size='32sp', bold=True, color=(1, 1, 1, 1), size_hint=(1, 0.1))
        self.layout.add_widget(title)

        # Add dropdown for selecting site
        self.site_dropdown = DropDown()
        for site in ["SITE1 (2038)", "SITE2 (2030)", "SITE3 (2028)", "SITE4 (2035)"]:
            btn = Button(text=site, size_hint_y=None, height=44)
            btn.bind(on_release=lambda btn: self.site_dropdown.select(btn.text))
            self.site_dropdown.add_widget(btn)

        self.site_button = Button(text='Select Site', size_hint=(1, 0.1))
        self.site_button.bind(on_release=self.site_dropdown.open)
        self.site_dropdown.bind(on_select=lambda instance, x: setattr(self.site_button, 'text', x))
        self.layout.add_widget(self.site_button)

        # Add buttons for questions
        questions_layout = GridLayout(cols=2, padding=10, spacing=10, size_hint=(1, 0.4))

        self.guidance_interrupted_button = Button(text='Was Guidance Interrupted?', size_hint=(1, None), height=50)
        self.guidance_interrupted_button.original_text = self.guidance_interrupted_button.text
        self.guidance_interrupted_button.bind(on_press=self.toggle_button)
        questions_layout.add_widget(self.guidance_interrupted_button)

        self.cat1_guidance_button = Button(text='Can the TLS still Provide CAT 1 Guidance?', size_hint=(1, None), height=50)
        self.cat1_guidance_button.original_text = self.cat1_guidance_button.text
        self.cat1_guidance_button.bind(on_press=self.toggle_button)
        questions_layout.add_widget(self.cat1_guidance_button)

        self.component_replacement_button = Button(text='Was a Major Component Replacement Required?', size_hint=(1, None), height=50)
        self.component_replacement_button.original_text = self.component_replacement_button.text
        self.component_replacement_button.bind(on_press=self.toggle_button)
        questions_layout.add_widget(self.component_replacement_button)

        self.provide_archive_button = Button(text='Can you Provide an Archive?', size_hint=(1, None), height=50)
        self.provide_archive_button.original_text = self.provide_archive_button.text
        self.provide_archive_button.bind(on_press=self.toggle_button)
        questions_layout.add_widget(self.provide_archive_button)

        # Add Alert/Alarm Code Present button
        self.alert_code_button = Button(text='Alert/Alarm Code Present?', size_hint=(1, None), height=50)
        self.alert_code_button.original_text = self.alert_code_button.text
        self.alert_code_button.bind(on_press=self.toggle_alert_code_button)
        questions_layout.add_widget(self.alert_code_button)

        # Add FMC/PMC/NMC dropdown button
        self.status_dropdown = DropDown()
        for status, color, tooltip_text in [("FMC", (0, 1, 0, 1), "TLS is fully operational with no Alerts or Alarms Present."),
                                            ("PMC", (1, 1, 0, 1), "TLS is able to provide CAT 1 Guidance, but there is a problem that needs to be addressed before it evolves."),
                                            ("NMC", (1, 0, 0, 1), "TLS is UNABLE to Provide CAT 1 Guidance.")]:
            btn = Button(text=status, size_hint_y=None, height=44, background_color=color)
            btn.tooltip_text = tooltip_text
            btn.bind(on_release=lambda btn: self.status_dropdown.select(btn.text))
            self.status_dropdown.add_widget(btn)

        self.status_button = Button(text='Select Status', size_hint=(1, None), height=50)
        self.status_button.bind(on_release=self.status_dropdown.open)
        self.status_dropdown.bind(on_select=self.update_status_button)
        questions_layout.add_widget(self.status_button)

        self.layout.add_widget(questions_layout)

        # Add comments section
        self.comments_input = TextInput(hint_text='Comments', size_hint=(1, 0.2))
        self.layout.add_widget(self.comments_input)

        # Add ARCHIVE button
        self.archive_button = Button(text='ARCHIVE', size_hint=(1, 0.1))
        self.archive_button.bind(on_press=self.select_archive_folder)
        self.layout.add_widget(self.archive_button)

        # Add Connect & Submit button
        self.connect_button = Button(text='Connect & Submit', size_hint=(1, 0.1), background_color=(1, 0, 0, 1))  # Red color
        self.connect_button.bind(on_press=self.connect_and_submit)
        self.layout.add_widget(self.connect_button)

        # Bind mouse position to show tooltips
        Window.bind(mouse_pos=self.on_mouse_pos)

        return self.layout

    def _update_rect(self, instance, value):
        self.rect.pos = instance.pos
        self.rect.size = instance.size

    def toggle_button(self, instance):
        if instance.text == 'YES':
            instance.text = instance.original_text
        else:
            instance.text = 'YES'

    def toggle_alert_code_button(self, instance):
        if instance.text == 'YES':
            instance.text = instance.original_text
            self.layout.remove_widget(self.alert_code_input)
            del self.alert_code_input
        else:
            instance.text = 'YES'
            self.alert_code_input = TextInput(size_hint=(1, None), height=50)
            self.layout.add_widget(self.alert_code_input, index=self.layout.children.index(self.comments_input) + 1)

    def show_tooltip(self, title, text, pos):
        if hasattr(self, 'tooltip'):
            self.tooltip.dismiss()
            del self.tooltip
        self.tooltip = Tooltip(title=title, text=text)
        self.tooltip.open()
        self.tooltip.pos = pos

    def hide_tooltip(self):
        if hasattr(self, 'tooltip'):
            self.tooltip.dismiss()
            del self.tooltip

    def on_mouse_pos(self, *args):
        pos = args[1]
        if self.status_dropdown.attach_to:
            for btn in self.status_dropdown.container.children:
                if btn.collide_point(*btn.to_widget(*pos)):
                    self.show_tooltip(btn.text, btn.tooltip_text, pos)
                    return
        self.hide_tooltip()

    def update_status_button(self, instance, text):
        self.status_button.text = text
        if text == "FMC":
            self.status_button.background_color = (0, 1, 0, 1)  # Green
        elif text == "PMC":
            self.status_button.background_color = (1, 1, 0, 1)  # Yellow
        elif text == "NMC":
            self.status_button.background_color = (1, 0, 0, 1)  # Red

    def select_archive_folder(self, instance):
        Tk().withdraw()  # Hide the root window
        self.archive_folder = askdirectory()
        if self.archive_folder:
            instance.text = 'Archive Selected'

    def is_connected(self):
        try:
            # Connect to the host -- tells us if the host is actually reachable
            socket.create_connection(("www.google.com", 80))
            return True
        except OSError:
            pass
        return False

    def connect_and_submit(self, instance):
        # Functionality for the Connect & Submit button
        if not self.is_connected():
            instance.background_color = (1, 0, 0, 1)  # Red color
            logging.error("No internet connection.")
            return

        try:
            # Authenticate and create the Gmail API service
            creds = None
            BASE_DIR = os.path.dirname(os.path.abspath(__file__))
            CREDENTIALS_PATH = os.path.join(BASE_DIR, 'credentials.json')
            TOKEN_PATH = os.path.join(BASE_DIR, 'token.json')

            if os.path.exists(TOKEN_PATH):
                creds = Credentials.from_authorized_user_file(TOKEN_PATH, SCOPES)
            if not creds or not creds.valid:
                if creds and creds.expired and creds.refresh_token:
                    creds.refresh(Request())
                else:
                    flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_PATH, SCOPES)
                    creds = flow.run_local_server(port=0)
                with open(TOKEN_PATH, 'w') as token:
                    token.write(creds.to_json())

            service = build('gmail', 'v1', credentials=creds)

            sender_email = "anpctechs@gmail.com"
            recipients = ["jtester@anpc.com", "ibeckenbach@anpc.com", "kwinner@anpc.com"]
            subject = "New Support Ticket"
            site = self.site_button.text
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            guidance_interrupted = "YES" if self.guidance_interrupted_button.text == 'YES' else "NO"
            cat1_guidance = "YES" if self.cat1_guidance_button.text == 'YES' else "NO"
            component_replacement = "YES" if self.component_replacement_button.text == 'YES' else "NO"
            provide_archive = "YES" if self.provide_archive_button.text == 'YES' else "NO"
            alert_code = self.alert_code_input.text if hasattr(self, 'alert_code_input') else ''
            status = self.status_button.text
            archive_folder = self.archive_folder if hasattr(self, 'archive_folder') else ''
            comments = self.comments_input.text

            data = {
                'Site': site,
                'Timestamp': timestamp,
                'Guidance Interrupted': guidance_interrupted,
                'CAT 1 Guidance': cat1_guidance,
                'Component Replacement': component_replacement,
                'Provide Archive': provide_archive,
                'Alert Code': alert_code,
                'Status': status,
                'Archive Folder': archive_folder,
                'Comments': comments
            }

            # Create a DataFrame and save it to an Excel file
            df = pd.DataFrame([data])
            excel_file_path = os.path.join(os.path.expanduser("~"), "Desktop", 'support_ticket.xlsx')
            try:
                df.to_excel(excel_file_path, index=False)
            except PermissionError:
                messagebox.showerror("Permission Denied", f"Permission denied: '{excel_file_path}'. Please close the file if it is open and try again.")
                return

            # Set font size to 12 points in the Excel file
            wb = Workbook()
            ws = wb.active

            # Apply styles to the header
            header_font = Font(bold=True, size=14, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            for col_num, column_title in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_num, value=column_title)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Write data to the Excel file
            for row_num, row_data in enumerate(df.values, 2):
                for col_num, cell_value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=cell_value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # Adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

            wb.save(excel_file_path)

            # Zip the contents of the selected folder
            if archive_folder:
                zip_file_path = os.path.join(os.path.expanduser("~"), "Desktop", 'archive.zip')
                try:
                    with zipfile.ZipFile(zip_file_path, 'w') as zipf:
                        for root, dirs, files in os.walk(archive_folder):
                            for file in files:
                                zipf.write(os.path.join(root, file), os.path.relpath(os.path.join(root, file), archive_folder))
                except PermissionError:
                    messagebox.showerror("Permission Denied", f"Permission denied: '{zip_file_path}'. Please close the file if it is open and try again.")
                    return

            # Email body
            body = f"""
            <html>
            <body>
            <p>Site: {site}</p>
            <p>Timestamp: {timestamp}</p>
            <p>Guidance Interrupted: {guidance_interrupted}</p>
            <p>CAT 1 Guidance: {cat1_guidance}</p>
            <p>Component Replacement: {component_replacement}</p>
            <p>Provide Archive: {provide_archive}</p>
            <p>Alert Code: {alert_code}</p>
            <p>Status: {status}</p>
            <p>Archive Folder: {archive_folder}</p>
            <p>Comments: {comments}</p>
            </body>
            </html>
            """

            # Send email
            send_email(service, sender_email, recipients, subject, body, attachments=[excel_file_path, zip_file_path] if archive_folder else [excel_file_path])

            # Change Connect & Submit button color to green
            instance.background_color = (0, 1, 0, 1)  # Green color

            # Close the application
            App.get_running_app().stop()
        except Exception as e:
            logging.error(f"Error connecting to Gmail: {e}")

if __name__ == '__main__':
    try:
        ANPCSupportTicket().run()
    except Exception as e:
        logging.error(f"Error running app: {e}")
